def openmyxlsx(xlxs_name):
	import openpyxl
	wb=openpyxl.load_workbook(xlxs_name)
	a=wb.active
	sendinfo=[]
	for x in range(1,len(a['A'])+1):
		sendinfo.append([a.cell(row=x,column=1).value,a.cell(row=x,column=2).value,a.cell(row=x,column=3).value])
	return sendinfo
